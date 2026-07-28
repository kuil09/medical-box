import csv
import hashlib
import io
import re
import zipfile
from collections.abc import Iterator
from itertools import chain, islice
from pathlib import PurePosixPath
from typing import Any
from urllib.parse import unquote

import httpx
from openpyxl import load_workbook
from tenacity import retry, retry_if_exception_type, stop_after_attempt, wait_exponential

from .sources import SourceDefinition


class SourceResponseError(RuntimeError):
    pass


class SourceTotalChangedError(SourceResponseError):
    pass


MAX_SOURCE_FILE_BYTES = 250 * 1024 * 1024
SOURCE_API_TIMEOUT_SECONDS = 30
SOURCE_FILE_TIMEOUT_SECONDS = 300
SOURCE_FILE_HEAD_TIMEOUT_SECONDS = 15


def unwrap_item_container(value: Any) -> Any:
    while isinstance(value, dict) and len(value) == 1 and "item" in value:
        value = value["item"]
    return value


class PublicDataFetcher:
    def __init__(self, service_key: str, page_size: int = 500) -> None:
        self.service_key = service_key
        self.page_size = page_size

    def pages(self, source: SourceDefinition) -> Iterator[tuple[int, list[dict[str, Any]], int]]:
        if not source.api_url:
            return
        page = 1
        expected_total: int | None = None
        while True:
            items, total = self._get_items_page(source.api_url, page)
            if expected_total is None:
                expected_total = total
            elif total != expected_total:
                raise SourceTotalChangedError(
                    f"Source total changed during pagination: {expected_total} to {total}."
                )
            if not items and (page - 1) * self.page_size < total:
                raise SourceResponseError(
                    f"Source ended at page {page} before {total} records were collected."
                )
            yield page, items, total
            if page * self.page_size >= total:
                return
            page += 1

    def tabular_file(
        self,
        source: SourceDefinition,
    ) -> tuple[str, str | None, list[dict[str, Any]]]:
        if not source.api_url:
            raise SourceResponseError("Tabular source URL is not configured.")
        response = self._get_file(source.api_url)
        content = response.content
        records = self._extract_tabular_records(
            content,
            filename=PurePosixPath(response.url.path).name,
            content_type=response.headers.get("content-type", ""),
        )
        if not records:
            raise SourceResponseError("Tabular source did not contain data rows.")
        return (
            hashlib.sha256(content).hexdigest(),
            self._file_version_marker(response),
            records,
        )

    def tabular_file_pages(
        self,
        source: SourceDefinition,
    ) -> tuple[
        str,
        str | None,
        Iterator[tuple[int, list[dict[str, Any]], int]],
    ]:
        if not source.api_url:
            raise SourceResponseError("Tabular source URL is not configured.")
        response = self._get_file(source.api_url)
        content = response.content
        filename = self._response_filename(response)
        content_type = response.headers.get("content-type", "")
        content_hash = hashlib.sha256(content).hexdigest()
        file_version = self._file_version_marker(response)
        if (
            not zipfile.is_zipfile(io.BytesIO(content))
            and PurePosixPath(filename).suffix.casefold() != ".xlsx"
            and "spreadsheetml" not in content_type.casefold()
        ):
            pages = self._csv_pages(content, self.page_size)
        else:
            records = self._extract_tabular_records(
                content,
                filename=filename,
                content_type=content_type,
            )
            if not records:
                raise SourceResponseError("Tabular source did not contain data rows.")
            pages = self._record_pages(records, self.page_size)
        return content_hash, file_version, pages

    def tabular_file_version(self, source: SourceDefinition) -> str | None:
        if not source.api_url:
            raise SourceResponseError("Tabular source URL is not configured.")
        try:
            response = self._head_file(source.api_url)
        except (httpx.HTTPError, SourceResponseError):
            return None
        return self._file_version_marker(response)

    @staticmethod
    def _content_disposition_filename(response: httpx.Response) -> str | None:
        disposition = response.headers.get("content-disposition", "")
        match = re.search(
            r"filename\*?=(?:UTF-8''|\"?)([^\";]+)",
            disposition,
            flags=re.IGNORECASE,
        )
        if match:
            return unquote(match.group(1).strip())
        return None

    @classmethod
    def _response_filename(cls, response: httpx.Response) -> str:
        return cls._content_disposition_filename(response) or PurePosixPath(
            response.url.path
        ).name

    @classmethod
    def _file_version_marker(cls, response: httpx.Response) -> str | None:
        filename = cls._content_disposition_filename(response)
        raw_content_length = response.headers.get("content-length")
        try:
            content_length = int(raw_content_length or "0")
        except ValueError:
            return None
        if not filename or content_length <= 0:
            return None
        etag = response.headers.get("etag")
        last_modified = response.headers.get("last-modified")
        identity_parts = [
            f"filename={filename}",
            f"bytes={content_length}",
            f"etag={etag}" if etag else "",
            f"last-modified={last_modified}" if last_modified else "",
        ]
        identity = "|".join(part for part in identity_parts if part)
        if not identity:
            return None
        return f"file-version:{hashlib.sha256(identity.encode()).hexdigest()}"

    @staticmethod
    def _record_pages(
        records: list[dict[str, Any]],
        page_size: int,
    ) -> Iterator[tuple[int, list[dict[str, Any]], int]]:
        total = len(records)
        for offset in range(0, total, page_size):
            yield offset // page_size + 1, records[offset : offset + page_size], total

    @retry(
        retry=retry_if_exception_type((httpx.HTTPError, SourceResponseError)),
        stop=stop_after_attempt(2),
        wait=wait_exponential(multiplier=0.5, min=0.5, max=8),
        reraise=True,
    )
    def _head_file(self, url: str) -> httpx.Response:
        response = httpx.head(
            url,
            timeout=SOURCE_FILE_HEAD_TIMEOUT_SECONDS,
            follow_redirects=True,
        )
        response.raise_for_status()
        self._validate_declared_file_size(response)
        return response

    @retry(
        retry=retry_if_exception_type((httpx.HTTPError, SourceResponseError)),
        stop=stop_after_attempt(4),
        wait=wait_exponential(multiplier=0.5, min=0.5, max=8),
        reraise=True,
    )
    def _get_file(self, url: str) -> httpx.Response:
        response = httpx.get(
            url,
            timeout=httpx.Timeout(
                SOURCE_FILE_TIMEOUT_SECONDS,
                connect=SOURCE_API_TIMEOUT_SECONDS,
                write=SOURCE_API_TIMEOUT_SECONDS,
                pool=SOURCE_API_TIMEOUT_SECONDS,
            ),
            follow_redirects=True,
        )
        response.raise_for_status()
        self._validate_declared_file_size(response)
        declared_size = int(response.headers.get("content-length", "0") or "0")
        if declared_size > MAX_SOURCE_FILE_BYTES or len(response.content) > MAX_SOURCE_FILE_BYTES:
            raise SourceResponseError("Tabular source exceeds the 250 MiB safety limit.")
        return response

    @staticmethod
    def _validate_declared_file_size(response: httpx.Response) -> None:
        raw_size = response.headers.get("content-length", "0") or "0"
        try:
            declared_size = int(raw_size)
        except ValueError as exc:
            raise SourceResponseError(
                "Tabular source returned an invalid content length."
            ) from exc
        if declared_size > MAX_SOURCE_FILE_BYTES:
            raise SourceResponseError("Tabular source exceeds the 250 MiB safety limit.")

    @retry(
        retry=retry_if_exception_type(SourceResponseError),
        stop=stop_after_attempt(4),
        wait=wait_exponential(multiplier=0.5, min=0.5, max=8),
        reraise=True,
    )
    def _get_items_page(
        self,
        url: str,
        page: int,
    ) -> tuple[list[dict[str, Any]], int]:
        return self._extract_items(self._get_json(url, page))

    @retry(
        retry=retry_if_exception_type(httpx.HTTPError),
        stop=stop_after_attempt(4),
        wait=wait_exponential(multiplier=0.5, min=0.5, max=8),
        reraise=True,
    )
    def _get_json(self, url: str, page: int) -> dict[str, Any]:
        response = httpx.get(
            url,
            params={
                "serviceKey": self.service_key,
                "pageNo": page,
                "numOfRows": self.page_size,
                "type": "json",
            },
            timeout=SOURCE_API_TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        try:
            payload = response.json()
        except ValueError as exc:
            raise SourceResponseError("Source returned non-JSON content.") from exc
        if not isinstance(payload, dict):
            raise SourceResponseError("Source returned an unexpected top-level value.")
        return payload

    @staticmethod
    def _extract_items(payload: dict[str, Any]) -> tuple[list[dict[str, Any]], int]:
        response = payload.get("response") if isinstance(payload.get("response"), dict) else payload
        header = response.get("header") if isinstance(response, dict) else None
        if isinstance(header, dict):
            result_code = str(header.get("resultCode", "00"))
            if result_code not in {"00", "0", "NORMAL SERVICE"}:
                raise SourceResponseError(
                    f"Source error {result_code}: {header.get('resultMsg', 'unknown')}"
                )
        body = response.get("body", response) if isinstance(response, dict) else {}
        if not isinstance(body, dict):
            raise SourceResponseError("Source response body is not an object.")
        if "items" not in body:
            raise SourceResponseError("Source response is missing the items field.")
        raw_items = unwrap_item_container(body.get("items", []))
        if isinstance(raw_items, dict):
            items = [unwrap_item_container(raw_items)]
        elif isinstance(raw_items, list):
            items = [
                unwrapped
                for item in raw_items
                if isinstance(unwrapped := unwrap_item_container(item), dict)
            ]
        else:
            items = []
        try:
            total = int(body.get("totalCount") or len(items))
        except (TypeError, ValueError) as exc:
            raise SourceResponseError("Source response has an invalid total count.") from exc
        return items, total

    @staticmethod
    def _extract_tabular_records(
        content: bytes,
        *,
        filename: str,
        content_type: str,
    ) -> list[dict[str, Any]]:
        if zipfile.is_zipfile(io.BytesIO(content)):
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                names = [name for name in archive.namelist() if not name.endswith("/")]
                if "[Content_Types].xml" in names and any(
                    name.startswith("xl/") for name in names
                ):
                    return PublicDataFetcher._xlsx_records(content)
                candidates = [
                    name
                    for name in names
                    if PurePosixPath(name).suffix.casefold() in {".csv", ".xlsx"}
                ]
                if not candidates:
                    raise SourceResponseError("Archive has no supported CSV or XLSX file.")
                selected = sorted(candidates, key=lambda name: (name.count("/"), name))[0]
                nested = archive.read(selected)
                return PublicDataFetcher._extract_tabular_records(
                    nested,
                    filename=selected,
                    content_type="",
                )

        suffix = PurePosixPath(filename).suffix.casefold()
        if suffix == ".xlsx" or "spreadsheetml" in content_type.casefold():
            return PublicDataFetcher._xlsx_records(content)
        if suffix in {".xls", ".xlsb"}:
            raise SourceResponseError(
                "Legacy XLS/XLSB is unsupported; configure a CSV or XLSX source URL."
            )
        return PublicDataFetcher._csv_records(content)

    @staticmethod
    def _xlsx_records(content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        records: list[dict[str, Any]] = []
        try:
            for sheet in workbook.worksheets:
                rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
                records.extend(PublicDataFetcher._rows_to_records(rows))
        finally:
            workbook.close()
        return records

    @staticmethod
    def _csv_records(content: bytes) -> list[dict[str, Any]]:
        return [
            record
            for _, records, _ in PublicDataFetcher._csv_pages(content, 1_000)
            for record in records
        ]

    @staticmethod
    def _csv_pages(
        content: bytes,
        page_size: int,
    ) -> Iterator[tuple[int, list[dict[str, Any]], int]]:
        encoding: str | None = None
        for candidate in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                content[: 1024 * 1024].decode(candidate)
                encoding = candidate
                break
            except UnicodeDecodeError:
                continue
        if encoding is None:
            raise SourceResponseError("CSV encoding is not UTF-8, CP949, or EUC-KR.")
        text_stream = io.TextIOWrapper(
            io.BytesIO(content),
            encoding=encoding,
            newline="",
        )
        sample = text_stream.read(8192)
        text_stream.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            reader = csv.reader(text_stream, dialect)
        except csv.Error:
            delimiter = max((",", "\t", ";", "|"), key=sample.count)
            if sample.count(delimiter) == 0:
                raise SourceResponseError("CSV delimiter could not be identified.") from None
            reader = csv.reader(text_stream, delimiter=delimiter)
        prefix = [tuple(row) for row in islice(reader, 30)]
        header_index = PublicDataFetcher._header_index(prefix)
        if header_index is None:
            raise SourceResponseError("Tabular source header row could not be identified.")
        headers = PublicDataFetcher._unique_headers(prefix[header_index])
        records: list[dict[str, Any]] = []
        page = 1
        rows = chain(prefix[header_index + 1 :], (tuple(row) for row in reader))
        for row in rows:
            values = list(row) + [None] * max(0, len(headers) - len(row))
            record = {
                header: value
                for header, value in zip(headers, values, strict=False)
                if value not in (None, "")
            }
            if not record:
                continue
            records.append(record)
            if len(records) == page_size:
                yield page, records, 0
                records = []
                page += 1
        if records:
            yield page, records, 0

    @staticmethod
    def _rows_to_records(rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
        header_index = PublicDataFetcher._header_index(rows)
        if header_index is None:
            raise SourceResponseError("Tabular source header row could not be identified.")
        headers = PublicDataFetcher._unique_headers(rows[header_index])
        records: list[dict[str, Any]] = []
        for row in rows[header_index + 1 :]:
            values = list(row) + [None] * max(0, len(headers) - len(row))
            record = {
                header: value
                for header, value in zip(headers, values, strict=False)
                if value not in (None, "")
            }
            if record:
                records.append(record)
        return records

    @staticmethod
    def _header_index(rows: list[tuple[Any, ...]]) -> int | None:
        known_headers = {
            "표준코드",
            "품목기준코드",
            "제품명",
            "품목명",
            "업체명",
            "standardcode",
            "itemseq",
        }
        fallback: int | None = None
        for index, row in enumerate(rows[:30]):
            values = [str(value).strip() for value in row if value not in (None, "")]
            if len(values) < 2:
                continue
            fallback = index if fallback is None else fallback
            normalized = {value.replace(" ", "").casefold() for value in values}
            if normalized & known_headers:
                return index
        return fallback

    @staticmethod
    def _unique_headers(row: tuple[Any, ...]) -> list[str]:
        counts: dict[str, int] = {}
        headers: list[str] = []
        for index, value in enumerate(row):
            base = str(value).strip() if value not in (None, "") else f"column_{index + 1}"
            count = counts.get(base, 0) + 1
            counts[base] = count
            headers.append(base if count == 1 else f"{base}_{count}")
        return headers
